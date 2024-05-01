
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)

wb=xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1']= 'No.'
ws['A1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)
ws['B1']= 'Movie Title'
ws['B1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)
ws['C1']= 'Release Date'
ws['C1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)
ws['D1']= 'Total Gross'
ws['D1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)
ws['E1']= 'Theaters'
ws['E1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)
ws['F1']= 'Average per theater'
ws['F1'].font = Font(name= 'Times New Roman',size=24, italic=False,bold=True)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 29
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 40


##SCARPING
movie_data = soup.findAll("td")
		
#print(movie_data[12].text)


counter = 1
for x in range(5):
    number = movie_data[counter-1].text
    name = movie_data[counter].text
    gross = float(movie_data[counter +4].text.strip('$').replace(',',''))
    theaters = float(movie_data[counter +5].text.replace(',',''))
    total_gross = float(movie_data[counter +6].text.strip('$').replace(',',''))
    release_date = movie_data[counter+7].text

    print(f'No.{number}')
    print(f'Movie name: {name}')
    print(f'Gross: {gross}')
    print(f'theaters: {theaters}')
    print(f'Total Gross: {total_gross}')
    print(f'Release Date: {release_date}')

    counter += 11





wb.save('Movie_Report.xlsx')